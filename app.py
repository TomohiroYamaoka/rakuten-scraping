from flask import Flask, url_for, render_template,flash,request, redirect
from bs4 import BeautifulSoup
#外部のサイトにアクセスすることができるモジュール
import requests,bs4
import os
#エクセルファイルをしようすることができるモジュール
from openpyxl import Workbook 

app = Flask(__name__)

#ページを開いた時に処理をするところ。
@app.route("/", methods=["GET"])
def index():
    return render_template("top.html")


@app.route("/input", methods=["GET","POST"])
def input(): 
    if request.form["inputText"]:
        #def fetch（）の処理を実行し、「result」に代入する。
        result=fetch()

        #結果を「result.html」に表示させる。
        return render_template("result.html",data=result)

    else:
        flash("URLを入力してください。")
        return render_template("top.html")    


def fetch():
    #入力したURLをurl代入する。
    url=request.form["inputText"]
    #楽天のwebページを取得する。
    response = requests.get(url)
    #resp=requests.get(address)
    html=response.content
    soup=BeautifulSoup(html,'html.parser')
    
    #空白の配列を作成する。
    data=[]
    
    
    #各項目のHTMLを取得する、
    #商品の使い道、商品を使う人
    contentA=soup.find_all("span",class_="revDispListTxt")
    #購入した回数
    contentB=soup.find_all("span",class_="revUserDispList")
    #本文(brを抜く)
    contentC=soup.find_all("dd",class_="revRvwUserEntryCmt description")
    #タイトル
    contentD=soup.find_all("span",class_="revUserRvwerNum value")
    #評価（星の数）

    #日付
    contentE=soup.find_all("span",class_="revUserEntryDate dtreviewed")

    #https://detail.chiebukuro.yahoo.co.jp/qa/question_detail/q11177607676
    #https://qiita.com/Gen6/items/eef1ac5f6b8bb816f677
    #https://qiita.com/toshiyuki_tsutsui/items/f143946944a428ed105b

    #HTMLを文字列にしてタグ、文字列を排除する。
    #https://qiita.com/hidetoshi_n_cograph/items/fca19b490fd0210e3633
    return contentA,contentB,contentC,contentD,contentE

    """
    wb=Workbook()
    ws=wb.activate

    #top.htmlでタイトルを出力させ、シート名になる
    ws.title='hoge'


    ws.title='Sample'
    ws['A1'].value='購入者名'
    ws['B1'].value='星の数'
    ws['C1'].value='商品の使い道'
    ws['D1'].value='商品を使う人'
    ws['E1'].value='購入した回数'
    ws['F1'].value='日付'
    ws['G1'].value='タイトル'
    ws['H1'].value='本文'
    
    #値を出力する。
    makeData()
    
    #saveでExcelファイルを保存する。
    wb.save('data.xlsx')
    """
if __name__ == '__main__':
    app.run(debug=True)